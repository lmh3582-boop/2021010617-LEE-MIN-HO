import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

def find_csv_files():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    print(f"Scanning directory: {base_dir}")
    
    output_file = os.path.join(base_dir, "CD_Summary.xlsx")
    all_data = []
    found_files = 0
    
    for root_dir, _, files in os.walk(base_dir):
        if "history_direct.csv" in files:
            csv_file = os.path.join(root_dir, "history_direct.csv")
            try:
                df = pd.read_csv(csv_file)
                
                # Find the CD column (case-insensitive and handles extra spaces/quotes)
                cd_columns = [col for col in df.columns if 'CD' in col.replace('"', '').replace(' ', '')]
                if cd_columns:
                    cd_col = cd_columns[0]
                    last_cd = df[cd_col].iloc[-1]
                    rel_path = os.path.relpath(root_dir, base_dir)
                    
                    all_data.append({
                        'Folder': os.path.basename(root_dir),
                        'Relative Path': rel_path,
                        'Last CD Value': last_cd,
                        'Total Rows': len(df),
                        'Full Path': root_dir
                    })
                    found_files += 1
                    print(f"Found CD value in {rel_path}: {last_cd}")
                else:
                    print(f"Warning: No CD column in {os.path.relpath(csv_file, base_dir)}")
                    print(f"Available columns: {df.columns.tolist()}")
            
            except Exception as e:
                print(f"Error processing {os.path.relpath(csv_file, base_dir)}: {str(e)}")
    
    if all_data:
        result_df = pd.DataFrame(all_data)
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            result_df.to_excel(writer, sheet_name='CD_Summary', index=False)
            workbook = writer.book
            worksheet = writer.sheets['CD_Summary']
            
            for i, col in enumerate(result_df.columns, 1):
                max_length = max((
                    result_df[col].astype(str).map(len).max(),
                    len(str(col))
                ))
                worksheet.column_dimensions[get_column_letter(i)].width = min(max_length + 2, 50)
            
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            print(f"\nSuccess! Found {found_files} files with CD values.")
            print(f"Summary saved to: {output_file}")
    else:
        print("\nNo history_direct.csv files with CD column found in the directory tree.")

if __name__ == "__main__":
    find_csv_files()